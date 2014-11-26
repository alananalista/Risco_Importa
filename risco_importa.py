# -*- coding: ascii -*-
import sys
import os
from openpyxl import load_workbook

os.system("cls")
arquivoBase = 'C:\\#BASEKB-ChangeHere.xlsx'
arquivoNovo = sys.argv[1]

print('******************************************************************')
print('[!] - Application Copy Hash between knowledge Base files')
print('[!] - Aplicativo de Copia de Hash entre Arquivos de Knowledge Base')
print('******************************************************************')
print('[0] - Opening file - ') + arquivoNovo 
wb_novo = load_workbook(sys.argv[1])
print('[0] - Opening source sheet')  
sheet_Novo = wb_novo['General Information']
print('[0] - Capturando hash origem') 
hash_Novo = sheet_Novo['C12'].value
print('[0] - Hash origem encontrado - ') + hash_Novo
print()
print()
print('[1] - Abrindo arquivo destino')
wbBase = load_workbook(arquivoBase)
#print wbBase.get_sheet_names()
print('[1] - Abrindo planilha destino ')
sheetBase = wbBase['General Information']
print('[1] - Gravando novo hash')
sheetBase['C12'] = hash_Novo
print('[1] - Hash novo gravado')
wbBase.save(arquivoBase)

print('[2] - Removendo arquivo lido')
os.rename(arquivoNovo,arquivoNovo +'-'+ hash_Novo + '.aaa')

print('******************************************************************')
print('[!] - Operacao Concluida com sucesso')
print('******************************************************************')

